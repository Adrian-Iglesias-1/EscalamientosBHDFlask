import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime

class ExcelHandler:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.data = {
            'contactos': {},
            'contactos_suc': {},
            'contactos_finde': {},
            'unificado': {}
        }

    def normalizar(self, s):
        if pd.isna(s):
            return ""
        t = str(s).upper().strip()
        for char in [".", "-", "_", " ", "/"]:
            t = t.replace(char, "")
        return t

    def normalizar_custodio(self, region, zone):
        zone = str(zone).strip().upper() if pd.notna(zone) else ''
        if 'Off-prime Brinks' in region: return f'Brinks {zone}'
        if 'Off-prime STE' in region: return 'BHD - STE Metro'
        if region == 'Sucursal-Sucursal': return 'SUCURSAL'
        if 'Sucursal-DriveUP Brinks' in region: return f'Brinks {zone}'
        if 'Sucursal-DriveUP STE' in region: return 'BHD - STE Metro'
        if 'Sucursal-Sucursal STE' in region: return 'BHD - STE Metro'
        if region in ['SUCURSAL', 'BHD - STE Metro', 'Brinks METRO', 'Brinks NORTE', 'Brinks ESTE', 'Brinks SUR']:
            return region
        return region

    def cargar_datos(self):
        if not os.path.exists(self.excel_path):
            return False, "Planilla no encontrada"
        
        try:
            xls = pd.ExcelFile(self.excel_path)
            
            # 1. CONTACTOS SEMANA
            if "CONTACTOS SEMANA" in xls.sheet_names:
                df_semana = pd.read_excel(xls, "CONTACTOS SEMANA")
                self.data['contactos'] = {
                    self.normalizar(r.iloc[0]): [r.iloc[1], r.iloc[2]]
                    for _, r in df_semana.iterrows() if not pd.isna(r.iloc[0])
                }

            # 2. CONTACTOS_SUC
            if "CONTACTOS_SUC" in xls.sheet_names:
                df_suc = pd.read_excel(xls, "CONTACTOS_SUC")
                for _, row in df_suc.iterrows():
                    if len(row) > 0 and pd.notna(row.iloc[0]):
                        sid = self.normalizar(row.iloc[0])
                        semail = row.iloc[6] if len(row) > 6 else ""
                        scopia = row.iloc[7] if len(row) > 7 else ""
                        self.data['contactos_suc'][sid] = [semail, scopia]

            # 3. CONTACTOS FINDE
            if "CONTACTOS FINDE" in xls.sheet_names:
                df_finde = pd.read_excel(xls, "CONTACTOS FINDE")
                for _, row in df_finde.iterrows():
                    reg = str(row.get('REGIONES', "")).strip()
                    if reg:
                        self.data['contactos_finde'][self.normalizar(reg)] = [row.get('CONTACTOS', ""), row.get('COPIA', "")]

            # 4. UNIFICADO
            if "UNIFICADO" in xls.sheet_names:
                df_unif = pd.read_excel(xls, "UNIFICADO")
                for _, row in df_unif.iterrows():
                    if pd.notna(row.iloc[0]):
                        id_n = self.normalizar(str(row.iloc[0]))
                        self.data['unificado'][id_n] = {
                            'nombre': str(row.iloc[1]) if len(row) > 1 else "",
                            'custodio': str(row.iloc[2]) if len(row) > 2 else "",
                            'sla_marcas': str(row.iloc[3]) if len(row) > 3 else "",
                            'sla_brinks': str(row.iloc[4]) if len(row) > 4 else "",
                            'denominacion': str(row.iloc[5]) if len(row) > 5 else "",
                            'zona': str(row.iloc[6]) if len(row) > 6 else "",
                            'disp_o_mult': str(row.iloc[7]) if len(row) > 7 else "",
                            'address2': str(row.iloc[8]) if len(row) > 8 else "",
                            'city': str(row.iloc[9]) if len(row) > 9 else "",
                            'ip_address': str(row.iloc[10]) if len(row) > 10 else "",
                            'district': str(row.iloc[11]) if len(row) > 11 else ""
                        }
            return True, "Datos cargados correctamente"
        except Exception as e:
            return False, str(e)

    def guardar_atm(self, atm_id, nombre, sla, custodio):
        try:
            wb = load_workbook(self.excel_path)
            # RCU sheet
            if "RCU" in wb.sheetnames:
                ws_rcu = wb["RCU"]
                new_row = ws_rcu.max_row + 1
                ws_rcu.cell(row=new_row, column=1, value=atm_id.upper())
                ws_rcu.cell(row=new_row, column=3, value=nombre)
                ws_rcu.cell(row=new_row, column=12, value=sla)
            # SLA sheet
            if "SLA" in wb.sheetnames:
                ws_sla = wb["SLA"]
                new_row = ws_sla.max_row + 1
                ws_sla.cell(row=new_row, column=1, value=atm_id.upper())
                ws_sla.cell(row=new_row, column=4, value=custodio)
                ws_sla.cell(row=new_row, column=8, value=sla)
            
            # Update UNIFICADO sheet as well
            if "UNIFICADO" in wb.sheetnames:
                ws_unif = wb["UNIFICADO"]
                new_row = ws_unif.max_row + 1
                ws_unif.cell(row=new_row, column=1, value=atm_id.upper())
                ws_unif.cell(row=new_row, column=2, value=nombre)
                ws_unif.cell(row=new_row, column=3, value=custodio)
                ws_unif.cell(row=new_row, column=4, value=sla)
                ws_unif.cell(row=new_row, column=12, value=sla)

            wb.save(self.excel_path)
            return True, "ATM guardado"
        except Exception as e:
            return False, str(e)

    def procesar_rcu(self, ruta_rcu_nuevo):
        """Procesa archivo RCU desde ruta en disco."""
        df_nuevo = pd.read_excel(ruta_rcu_nuevo, header=2)
        return self.procesar_rcu_desde_df(df_nuevo)

    def procesar_rcu_desde_df(self, df_nuevo):
        """Procesa RCU desde DataFrame (evita guardar a disco)."""
        try:
            wb = load_workbook(self.excel_path)
            
            # Ensure UNIFICADO exists
            if "UNIFICADO" not in wb.sheetnames:
                ws_unif = wb.create_sheet("UNIFICADO")
                headers = ["ID", "NOMBRE", "CUSTODIO", "SLA_MARCAS", "SLA_BRINKS", "DENOMINACION", "ZONA", "DISP_O_MULT", "ADDRESS2", "CITY", "IP_ADDRESS", "DISTRICT"]
                for i, h in enumerate(headers, 1): ws_unif.cell(row=1, column=i, value=h)
            else:
                ws_unif = wb["UNIFICADO"]

            # Map existing IDs
            ids_unif = {}
            for row in range(2, ws_unif.max_row + 1):
                cell_val = ws_unif.cell(row=row, column=1).value
                if cell_val: ids_unif[str(cell_val).strip().upper()] = row

            ws_rcu = wb["RCU"] if "RCU" in wb.sheetnames else None
            if not ws_rcu: return False, "No se encontró hoja RCU"

            ids_planilla = {}
            for row in range(2, ws_rcu.max_row + 1):
                cell_val = ws_rcu.cell(row=row, column=1).value
                if cell_val: ids_planilla[str(cell_val).strip().upper()] = row

            dict_sla_marcas = {}
            if "SLA" in wb.sheetnames:
                ws_sla = wb["SLA"]
                for row in range(2, ws_sla.max_row + 1):
                    id_sla = ws_sla.cell(row=row, column=1).value
                    sla_marca = ws_sla.cell(row=row, column=9).value
                    if id_sla and sla_marca: dict_sla_marcas[str(id_sla).strip().upper()] = str(sla_marca)

            actualizados = 0
            nuevos = 0
            contactos_limpiados = 0
            contactos_propagados = 0

            # Precargar mapa de emails por custodio para propagación
            email_por_custodio = {}
            if "CONTACTOS SEMANA" in wb.sheetnames:
                ws_cont = wb["CONTACTOS SEMANA"]
                for row in ws_cont.iter_rows(min_row=2):
                    key = str(row[0].value or '').strip()
                    if not key:
                        continue
                    key_norm = self.normalizar(key)
                    e = str(row[1].value or '').strip() if row[1].value else ''
                    c = str(row[2].value or '').strip() if row[2].value else ''
                    if e and e != 'nan':
                        # Si la key es un ATM ID, buscar su custodio
                        if key_norm in self.data['unificado']:
                            cust = self.data['unificado'][key_norm].get('custodio', '')
                            if cust and cust not in email_por_custodio:
                                email_por_custodio[cust] = (e, c)
                        # Si no, la key puede ser nombre de custodio
                        elif key_norm not in email_por_custodio:
                            email_por_custodio[key_norm] = (e, c)
            
            for _, row in df_nuevo.iterrows():
                id_raw = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                if not id_raw or id_raw.upper() in ["ID", "ID2", "ID3", "ID4"]: continue
                id_upper = id_raw.upper()
                
                # Extract fields
                address2 = str(row.get('ADDRESS2', '')) if pd.notna(row.get('ADDRESS2', None)) else ""
                region = str(row.get('REGION', '')) if pd.notna(row.get('REGION', None)) else ""
                zone = str(row.get('ZONE', '')) if pd.notna(row.get('ZONE', None)) else ""
                
                custodio_nuevo = self.normalizar_custodio(region, zone)
                sla_val = dict_sla_marcas.get(id_upper, "Sin SLA")

                # Guardar custodio VIEJO antes de actualizar
                custodio_viejo = ""
                if id_upper in ids_unif:
                    old_row = ids_unif[id_upper]
                    custodio_viejo = str(ws_unif.cell(row=old_row, column=3).value or '').strip()

                # ── Escribir TODOS los datos en RCU ──
                if id_upper in ids_planilla:
                    rc_row = ids_planilla[id_upper]
                    actualizados += 1
                else:
                    rc_row = ws_rcu.max_row + 1
                    ws_rcu.cell(row=rc_row, column=1, value=id_upper)
                    nuevos += 1
                    ids_planilla[id_upper] = rc_row

                # Mapeo completo RCU nuevo → planilla RCU
                # Col planilla: 1=ID2, 2=ADDRESS, 3=ADDRESS2, 4=CITY, 5=STATE, 6=OBJECT TYPE,
                # 7=BRANCH, 8=REGION, 9=ZONE, 10=AREA, 11=SECTOR, 12=DISTRICT,
                # 13=DATA LINE, 14=SERIAL NUM, 15=IP ADDRESS, 16=SITE CODE,
                # 17=PRODUCT REFERENCE, 18=MASTER CUSTOMER NUMBER
                mapeo_rcu = {
                    2: 'ADDRESS', 3: 'ADDRESS2', 4: 'CITY', 5: 'STATE',
                    6: 'OBJECT TYPE', 7: 'BRANCH', 8: 'REGION', 9: 'ZONE',
                    10: 'AREA', 11: 'SECTOR', 12: 'DISTRICT',
                    13: 'DATA LINE', 14: 'SERIAL NUM', 15: 'IP ADDRESS',
                    16: 'SITE CODE', 17: 'PRODUCT REFERENCE', 18: 'MASTER CUSTOMER NUMBER'
                }
                for col_idx, field in mapeo_rcu.items():
                    val = row.get(field, '')
                    if pd.notna(val):
                        ws_rcu.cell(row=rc_row, column=col_idx, value=str(val))

                # Update/Add to UNIFICADO
                if id_upper in ids_unif:
                    u_row = ids_unif[id_upper]
                else:
                    u_row = ws_unif.max_row + 1
                    ws_unif.cell(row=u_row, column=1, value=id_upper)
                    ids_unif[id_upper] = u_row

                # UNIFICADO: solo NOMBRE, CUSTODIO, SLA
                ws_unif.cell(row=u_row, column=2, value=address2)
                ws_unif.cell(row=u_row, column=3, value=custodio_nuevo)
                ws_unif.cell(row=u_row, column=4, value=sla_val)

                # ── LIMPIEZA Y PROPAGACIÓN DE CONTACTOS ──
                id_norm = self.normalizar(id_upper)
                era_suc = 'SUCURSAL' in custodio_viejo.upper() or custodio_viejo.upper().startswith('SUC')
                es_suc = 'SUCURSAL' in custodio_nuevo.upper() or custodio_nuevo.upper().startswith('SUC')

                if custodio_viejo and custodio_viejo != custodio_nuevo:
                    # Limpiar entrada vieja
                    if era_suc and "CONTACTOS_SUC" in wb.sheetnames:
                        ws_suc = wb["CONTACTOS_SUC"]
                        for r in range(2, ws_suc.max_row + 1):
                            if self.normalizar(str(ws_suc.cell(row=r, column=1).value or '')) == id_norm:
                                ws_suc.cell(row=r, column=6).value = None
                                ws_suc.cell(row=r, column=7).value = None
                                contactos_limpiados += 1
                                break
                    if not era_suc and "CONTACTOS SEMANA" in wb.sheetnames:
                        ws_cont = wb["CONTACTOS SEMANA"]
                        for r in range(2, ws_cont.max_row + 1):
                            if self.normalizar(str(ws_cont.cell(row=r, column=1).value or '')) == id_norm:
                                ws_cont.cell(row=r, column=1).value = None
                                ws_cont.cell(row=r, column=2).value = None
                                contactos_limpiados += 1
                                break

                # Propagar contacto del nuevo custodio si existe
                if not es_suc and custodio_nuevo in email_por_custodio:
                    e, c = email_por_custodio[custodio_nuevo]
                    if "CONTACTOS SEMANA" in wb.sheetnames:
                        ws_cont = wb["CONTACTOS SEMANA"]
                        # Ver si ya tiene entrada
                        ya_tiene = False
                        for r in range(2, ws_cont.max_row + 1):
                            if self.normalizar(str(ws_cont.cell(row=r, column=1).value or '')) == id_norm:
                                # Ya existe, actualizar si está vacío
                                if not (ws_cont.cell(row=r, column=2).value or '').strip():
                                    ws_cont.cell(row=r, column=2).value = e
                                    ws_cont.cell(row=r, column=3).value = c
                                    contactos_propagados += 1
                                ya_tiene = True
                                break
                        if not ya_tiene:
                            nr = ws_cont.max_row + 1
                            ws_cont.cell(row=nr, column=1, value=id_upper)
                            ws_cont.cell(row=nr, column=2, value=e)
                            ws_cont.cell(row=nr, column=3, value=c)
                            contactos_propagados += 1

            # ── LIMPIAR ATMs QUE YA NO ESTÁN EN EL RCU ──
            ids_nuevo = set()
            for _, row in df_nuevo.iterrows():
                id_raw = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                if id_raw and id_raw.upper() not in ["ID", "ID2", "ID3", "ID4"]:
                    ids_nuevo.add(id_raw.upper())

            eliminados_rcu = 0
            rows_to_delete = []
            for row in range(2, ws_rcu.max_row + 1):
                cell_val = str(ws_rcu.cell(row=row, column=1).value or '').strip().upper()
                if cell_val and cell_val not in ids_nuevo:
                    rows_to_delete.append(row)
                    eliminados_rcu += 1

            # Borrar de abajo para arriba para no desfasar índices
            for row in reversed(rows_to_delete):
                ws_rcu.delete_rows(row)

            # Limpiar UNIFICADO también
            for row in range(2, ws_unif.max_row + 1):
                cell_val = str(ws_unif.cell(row=row, column=1).value or '').strip().upper()
                if cell_val and cell_val not in ids_nuevo:
                    ws_unif.cell(row=row, column=2).value = None

            total_planilla = ws_rcu.max_row - 1  # menos header
            wb.save(self.excel_path)
            return True, {
                'actualizados': actualizados,
                'nuevos': nuevos,
                'total_procesados': actualizados + nuevos,
                'total_planilla': total_planilla,
                'eliminados': eliminados_rcu,
                'contactos_limpiados': contactos_limpiados,
                'contactos_propagados': contactos_propagados
            }
        except Exception as e:
            return False, str(e)

    # =================================================================
    # TAB CONTACTOS — Obtener y actualizar contactos por custodio
    # =================================================================

    def obtener_contactos_custodio(self):
        """Devuelve lista de custodios con sus emails actuales desde la planilla."""
        self.cargar_datos()
        try:
            contactos_terceros = []
            contactos_sucursales = []
            contactos_finde_sucursal = {}

            # ── SUCURSAL finde ──
            if "SUCURSAL" in self.data['contactos_finde']:
                mail, cc = self.data['contactos_finde']["SUCURSAL"]
                email_str = str(mail) if pd.notna(mail) else ""
                cc_str = str(cc) if pd.notna(cc) and str(cc) != "nan" else ""
                contactos_finde_sucursal = {"email": email_str, "cc": cc_str}

            # ── TERCEROS ──
            # Paso 1: construir mapa custodio → (email L-V, cc L-V) desde CONTACTOS SEMANA
            # PRIMERO: claves que son nombres de custodio (prioridad más alta)
            custodio_email_semana = {}
            for key, (email_cont, cc_cont) in self.data['contactos'].items():
                e = str(email_cont) if pd.notna(email_cont) and str(email_cont) not in ("", "nan") else ""
                if not e:
                    continue
                c = str(cc_cont) if pd.notna(cc_cont) and str(cc_cont) != "nan" else ""
                # Buscar si la clave normalizada coincide con algún custodio de UNIFICADO
                key_norm = self.normalizar(key)
                matched = False
                for aid, ainfo in self.data['unificado'].items():
                    cust_norm = self.normalizar(ainfo.get('custodio', ''))
                    if cust_norm == key_norm:
                        cust_original = ainfo.get('custodio', '').strip()
                        custodio_email_semana[cust_original] = (e, c)
                        matched = True
                        break
                if matched:
                    continue

            # DESPUÉS: claves que son ATM IDs (solo si el custodio no tiene entrada por nombre)
            for key, (email_cont, cc_cont) in self.data['contactos'].items():
                e = str(email_cont) if pd.notna(email_cont) and str(email_cont) not in ("", "nan") else ""
                if not e:
                    continue
                c = str(cc_cont) if pd.notna(cc_cont) and str(cc_cont) != "nan" else ""
                if key not in self.data['unificado']:
                    continue
                cust = self.data['unificado'][key].get('custodio', '').strip()
                if cust and cust not in custodio_email_semana:
                    custodio_email_semana[cust] = (e, c)

            # Paso 2: construir mapa custodio → (email finde, cc finde) desde CONTACTOS FINDE
            custodio_email_finde = {}
            for fkey, (f_email, f_cc) in self.data['contactos_finde'].items():
                fkey_norm = self.normalizar(fkey).upper().replace(" ", "")
                e = str(f_email) if pd.notna(f_email) and str(f_email) not in ("", "nan") else ""
                if e:
                    c = str(f_cc) if pd.notna(f_cc) and str(f_cc) != "nan" else ""
                    # Asociar por nombre de custodio
                    custodio_email_finde[fkey_norm] = (e, c)

            # Paso 3: listar custodios únicos desde UNIFICADO
            custodios_vistos = set()
            for id_norm, info in self.data['unificado'].items():
                custodio = info.get('custodio', '').strip()
                if not custodio or 'SUCURSAL' in custodio.upper() or custodio.upper().startswith('SUC'):
                    continue
                if custodio in custodios_vistos:
                    continue
                custodios_vistos.add(custodio)

                # Email de semana (buscar por nombre exacto y por nombre normalizado)
                email = ""
                cc = ""
                cust_norm = self.normalizar(custodio)
                # Intentar match exacto primero
                if custodio in custodio_email_semana:
                    email, cc = custodio_email_semana[custodio]
                # Intentar match normalizado
                elif cust_norm in custodio_email_semana:
                    email, cc = custodio_email_semana[cust_norm]

                # Email de finde (buscar por coincidencia normalizada)
                aplica_finde = False
                email_finde = ""
                cc_finde = ""
                cust_norm = self.normalizar(custodio).upper().replace(" ", "")
                if cust_norm in custodio_email_finde:
                    aplica_finde = True
                    email_finde, cc_finde = custodio_email_finde[cust_norm]

                contactos_terceros.append({
                    "custodio": custodio,
                    "email": email,
                    "cc": cc,
                    "aplica_finde": aplica_finde,
                    "email_finde": email_finde,
                    "cc_finde": cc_finde
                })

            # ── SUCURSALES: uno por uno ──
            for id_norm, info in self.data['unificado'].items():
                custodio = info.get('custodio', '')
                cust_up = custodio.upper()
                if not ('SUCURSAL' in cust_up or cust_up.startswith('SUC')):
                    continue
                email = ""
                cc = ""
                if id_norm in self.data['contactos_suc']:
                    email = str(self.data['contactos_suc'][id_norm][0]) if pd.notna(self.data['contactos_suc'][id_norm][0]) else ""
                    cc = str(self.data['contactos_suc'][id_norm][1]) if len(self.data['contactos_suc'][id_norm]) > 1 and pd.notna(self.data['contactos_suc'][id_norm][1]) else ""
                contactos_sucursales.append({
                    "id": id_norm,
                    "nombre": info.get('nombre', ''),
                    "email": email,
                    "cc": cc if cc not in ("", "nan") else ""
                })

            return {
                "terceros": contactos_terceros,
                "sucursales": contactos_sucursales,
                "sucursal_finde": contactos_finde_sucursal
            }
        except Exception as e:
            return {"error": str(e)}

    def actualizar_contactos_custodio(self, custodio, email, cc, aplica_finde, tipo, email_finde="", cc_finde="", solo=""):
        """
        Actualiza contactos según el tipo de custodio.
        tipo: 'tercero' | 'sucursal' | 'sucursal_finde'
        solo: 'semana' (solo CONTACTOS SEMANA) | 'finde' (solo CONTACTOS FINDE) | '' (ambos)
        """
        from openpyxl import load_workbook
        try:
            wb = load_workbook(self.excel_path)
            cambios = 0

            if tipo == 'tercero':
                # Actualizar CONTACTOS SEMANA
                if (not solo or solo == 'semana') and "CONTACTOS SEMANA" in wb.sheetnames:
                    ws = wb["CONTACTOS SEMANA"]
                    cust_norm = self.normalizar(custodio)
                    for row in ws.iter_rows(min_row=2):
                        id_val = str(row[0].value or '').strip()
                        if not id_val:
                            continue
                        id_norm = self.normalizar(id_val)
                        # Caso A: col 0 es un ATM ID → ver si su custodio coincide
                        if id_norm in self.data['unificado'] and self.data['unificado'][id_norm].get('custodio', '') == custodio:
                            row[1].value = email if email else None
                            row[2].value = cc if cc else None
                            cambios += 1
                        # Caso B: col 0 es el nombre del custodio directamente
                        elif id_norm == cust_norm:
                            row[1].value = email if email else None
                            row[2].value = cc if cc else None
                            cambios += 1

                # Actualizar CONTACTOS FINDE
                if (not solo or solo == 'finde') and "CONTACTOS FINDE" in wb.sheetnames:
                    ws = wb["CONTACTOS FINDE"]
                    cust_norm = self.normalizar(custodio)
                    for row in ws.iter_rows(min_row=2):
                        key = str(row[0].value or '').strip()
                        if self.normalizar(key) == cust_norm:
                            row[1].value = (email_finde or email) if (email_finde or email) else None
                            row[2].value = (cc_finde or cc) if (cc_finde or cc) else None
                            cambios += 1
                            break
                    # Si no existe el custodio en FINDE, agregar fila
                    if cambios == 0 or not any(self.normalizar(str(ws.cell(row=r, column=1).value or '')) == cust_norm for r in range(2, ws.max_row + 1)):
                        nueva = ws.max_row + 1
                        ws.cell(row=nueva, column=1, value=custodio)
                        ws.cell(row=nueva, column=2, value=(email_finde or email) if (email_finde or email) else None)
                        ws.cell(row=nueva, column=3, value=(cc_finde or cc) if (cc_finde or cc) else None)
                        cambios += 1

            elif tipo == 'sucursal':
                atm_id_for_suc = custodio  # custodio es el ID del ATM
                # Actualizar CONTACTOS_SUC
                if "CONTACTOS_SUC" in wb.sheetnames:
                    ws = wb["CONTACTOS_SUC"]
                    for row in ws.iter_rows(min_row=2):
                        obj = str(row[0].value or '').strip()
                        if self.normalizar(obj) == self.normalizar(atm_id_for_suc):
                            row[6].value = email
                            row[7].value = cc
                            cambios += 1
                            break
                # También actualizar CONTACTOS SEMANA si existe
                if "CONTACTOS SEMANA" in wb.sheetnames:
                    ws_sem = wb["CONTACTOS SEMANA"]
                    atm_norm = self.normalizar(atm_id_for_suc)
                    for row in ws_sem.iter_rows(min_row=2):
                        id_val = str(row[0].value or '').strip()
                        if self.normalizar(id_val) == atm_norm:
                            row[1].value = email if email else None
                            row[2].value = cc if cc else None
                            cambios += 1
                            break

            elif tipo == 'sucursal_finde':
                if "CONTACTOS FINDE" in wb.sheetnames:
                    ws = wb["CONTACTOS FINDE"]
                    for row in ws.iter_rows(min_row=2):
                        key = str(row[0].value or '').strip()
                        if self.normalizar(key) == self.normalizar("SUCURSAL"):
                            row[1].value = email
                            row[2].value = cc
                            cambios += 1
                            break

            wb.save(self.excel_path)
            return {"status": "success", "cambios": cambios}
        except Exception as e:
            return {"error": str(e)}
