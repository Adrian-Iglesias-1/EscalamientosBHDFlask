import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import os


class ClosedAndBlockHandler:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self._ensure_file()

    def _ensure_file(self):
        if not os.path.exists(self.excel_path):
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'ClosedAndBlock'
            ws.append(['ID', 'NOMBRE', 'CUSTODIO', 'FECHA_INGRESO', 'ESTADO', 'ASUNTO', 'REPORTADO_POR'])
            wb.save(self.excel_path)

    def _load(self):
        df = pd.read_excel(self.excel_path, sheet_name='ClosedAndBlock')
        return df

    def normalizar(self, s):
        if pd.isna(s):
            return ""
        t = str(s).upper().strip()
        for char in [".", "-", "_", " ", "/"]:
            t = t.replace(char, "")
        return t

    def limpiar_vencidos(self):
        try:
            df = self._load()
            if df.empty:
                return 0, 0

            cutoff = datetime.now() - timedelta(hours=48)
            df['FECHA_INGRESO'] = pd.to_datetime(df['FECHA_INGRESO'], errors='coerce')

            before = len(df)
            df = df[df['FECHA_INGRESO'] >= cutoff]
            removed = before - len(df)

            if removed > 0:
                wb = load_workbook(self.excel_path)
                ws = wb['ClosedAndBlock']
                ids_vigentes = set(self.normalizar(i) for i in df['ID'].tolist() if pd.notna(i))
                rows_to_delete = []
                for row_idx in range(2, ws.max_row + 1):
                    cell_val = ws.cell(row=row_idx, column=1).value
                    if cell_val and self.normalizar(str(cell_val)) not in ids_vigentes:
                        rows_to_delete.append(row_idx)

                for row_idx in reversed(rows_to_delete):
                    ws.delete_rows(row_idx)

                wb.save(self.excel_path)

            return removed, len(df)
        except Exception as e:
            return 0, 0

    def agregar_ids(self, ids_raw, excel_unificado, asunto, reportado_por):
        try:
            self.limpiar_vencidos()
            df = self._load()

            added = 0
            skipped = 0

            for id_raw in ids_raw:
                id_norm = self.normalizar(id_raw)
                if not id_norm or id_norm == 'NAN':
                    continue

                existing_ids = set(self.normalizar(i) for i in df['ID'].tolist() if pd.notna(i))
                if id_norm in existing_ids:
                    skipped += 1
                    continue

                info = excel_unificado.get(id_norm, {})
                nombre = info.get('nombre', '') or id_raw
                custodio = info.get('custodio', 'N/A')

                wb = load_workbook(self.excel_path)
                ws = wb['ClosedAndBlock']
                new_row = ws.max_row + 1
                ws.cell(row=new_row, column=1, value=id_norm)
                ws.cell(row=new_row, column=2, value=nombre)
                ws.cell(row=new_row, column=3, value=custodio)
                ws.cell(row=new_row, column=4, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                ws.cell(row=new_row, column=5, value='BLOQUEADO')
                ws.cell(row=new_row, column=6, value=asunto)
                ws.cell(row=new_row, column=7, value=reportado_por)
                wb.save(self.excel_path)
                added += 1

                new_row_df = pd.DataFrame([{
                    'ID': id_norm,
                    'NOMBRE': nombre,
                    'CUSTODIO': custodio,
                    'FECHA_INGRESO': datetime.now(),
                    'ESTADO': 'BLOQUEADO',
                    'ASUNTO': asunto,
                    'REPORTADO_POR': reportado_por
                }])
                if df.empty:
                    df = new_row_df
                else:
                    df = pd.concat([df, new_row_df], ignore_index=True)

            return {'added': added, 'skipped': skipped, 'total': len(df)}

        except Exception as e:
            return {'error': str(e)}

    def buscar_ids(self, ids_raw):
        try:
            df = self._load()
            if df.empty:
                return []

            df['ID_NORM'] = df['ID'].apply(self.normalizar)

            results = []
            seen_ids = set()
            for id_raw in ids_raw:
                id_norm = self.normalizar(id_raw)
                if not id_norm or id_norm == 'NAN':
                    continue
                if id_norm in seen_ids:
                    continue
                match = df[df['ID_NORM'] == id_norm]
                if not match.empty:
                    seen_ids.add(id_norm)
                    row = match.iloc[0]
                    fecha = row['FECHA_INGRESO']
                    if pd.notna(fecha):
                        try:
                            fecha_dt = pd.to_datetime(fecha)
                            horas = (datetime.now() - fecha_dt.replace(tzinfo=None)).total_seconds() / 3600
                            horas_str = f"{int(horas)}h"
                        except:
                            horas_str = 'N/A'
                    else:
                        horas_str = 'N/A'
                    results.append({
                        'id': id_norm,
                        'nombre': row['NOMBRE'] if pd.notna(row['NOMBRE']) else '',
                        'custodio': row['CUSTODIO'] if pd.notna(row['CUSTODIO']) else '',
                        'fecha': str(fecha)[:16] if pd.notna(fecha) else '',
                        'horas': horas_str,
                        'estado': row['ESTADO'] if pd.notna(row['ESTADO']) else 'BLOQUEADO',
                        'asunto': row['ASUNTO'] if pd.notna(row.get('ASUNTO')) else '',
                        'reportado_por': row['REPORTADO_POR'] if pd.notna(row.get('REPORTADO_POR')) else ''
                    })

            return results
        except Exception as e:
            return []

    def listar_todos(self):
        try:
            df = self._load()
            if df.empty:
                return []
            results = []
            for _, row in df.iterrows():
                fecha = row.get('FECHA_INGRESO', '')
                if pd.notna(fecha):
                    try:
                        fecha_dt = pd.to_datetime(fecha)
                        horas = (datetime.now() - fecha_dt.replace(tzinfo=None)).total_seconds() / 3600
                        horas_str = f"{int(horas)}h"
                    except:
                        horas_str = 'N/A'
                else:
                    horas_str = 'N/A'
                results.append({
                    'id': str(row['ID']) if pd.notna(row['ID']) else '',
                    'nombre': row['NOMBRE'] if pd.notna(row['NOMBRE']) else '',
                    'custodio': row['CUSTODIO'] if pd.notna(row['CUSTODIO']) else '',
                    'fecha': str(fecha)[:16] if pd.notna(fecha) else '',
                    'horas': horas_str,
                    'estado': row['ESTADO'] if pd.notna(row['ESTADO']) else 'BLOQUEADO',
                    'asunto': row['ASUNTO'] if pd.notna(row.get('ASUNTO')) else '',
                    'reportado_por': row['REPORTADO_POR'] if pd.notna(row.get('REPORTADO_POR')) else ''
                })
            return results
        except Exception as e:
            return []

    def eliminar_id(self, id_raw):
        try:
            id_norm = self.normalizar(id_raw)
            df = self._load()
            if df.empty:
                return False

            df['ID_NORM'] = df['ID'].apply(self.normalizar)
            df = df[df['ID_NORM'] != id_norm]

            wb = load_workbook(self.excel_path)
            ws = wb['ClosedAndBlock']
            rows_to_delete = []
            for row_idx in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx, column=1).value
                if cell_val and self.normalizar(str(cell_val)) == id_norm:
                    rows_to_delete.append(row_idx)
            for row_idx in reversed(rows_to_delete):
                ws.delete_rows(row_idx)
            wb.save(self.excel_path)
            return True
        except Exception as e:
            return False
